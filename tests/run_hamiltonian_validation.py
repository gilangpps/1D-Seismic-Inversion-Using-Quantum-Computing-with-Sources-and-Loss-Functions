"""
run_hamiltonian_validation.py
──────────────────────────────────────────────────────────────────────────────
Standalone script to validate Hamiltonian evolution against classical solver.

This answers RM2 (Research Method 2): Does the Hamiltonian constructed via
Schrödingerisation correctly reproduce classical wave physics?

OUTPUT:
    - Validation plots: figures/hamiltonian_validation_*.png
    - Excel sheet: data/<timestamp>/hamiltonian_validation.xlsx
    - JSON report: data/<timestamp>/hamiltonian_validation.json

USAGE:
    python run_hamiltonian_validation.py
──────────────────────────────────────────────────────────────────────────────
"""

import numpy as np
import matplotlib.pyplot as plt
import json
import openpyxl
from pathlib import Path
from datetime import datetime

from src.experiment.validate_hamiltonian import validate_hamiltonian_evolution
from src.distributions import raised_cosine
from src.wave import gaussian_source, set_source_peak
from src.constants import OUTPUT_DIR, FIGURE_DIR


def main():
    print("=" * 70)
    print("HAMILTONIAN VALIDATION: exp(-iHt) vs Classical PDE Solver")
    print("=" * 70)

    # ── Configuration (same as main experiment) ──────────────────────────
    nx = 7
    dx = 63.0  # meters
    dt = 0.005  # seconds
    steps = 40
    t_max = steps * dt

    # Medium properties
    mu_min, mu_max = 1e10, 4e10
    rho_min, rho_max = 2e3, 4e3

    x_grid = np.arange(nx) * dx
    mu_arr = raised_cosine(x_grid, mu_min, mu_max, period=nx * dx)
    rho_arr = raised_cosine(x_grid, rho_min, rho_max, period=nx * dx)

    # Initial conditions
    u0 = np.zeros(nx)
    center = nx // 2
    u0[center] = 1.0  # Pulse at center
    v0 = np.zeros(nx)

    # Source
    source_func = gaussian_source(
        center_idx=center + 1,  # +1 for boundary offset
        width_idx=2.0,
        amplitude=1e6,
        nx_total=nx + 2,
    )
    t0 = t_max / 3.0
    sigma_t = t_max / 12.0
    set_source_peak(source_func, t0, sigma_t)

    print(f"\nConfiguration:")
    print(f"  Grid: nx={nx}, dx={dx} m, dt={dt} s, steps={steps}")
    print(f"  μ range: [{mu_min:.2e}, {mu_max:.2e}] Pa")
    print(f"  ρ range: [{rho_min:.2e}, {rho_max:.2e}] kg/m³")
    print(f"  Source: Gaussian at t0={t0:.3f} s, σ={sigma_t:.3f} s")

    # ── Run validation ───────────────────────────────────────────────────
    print("\nRunning validation...")
    results = validate_hamiltonian_evolution(
        mu_arr=mu_arr,
        rho_arr=rho_arr,
        u0=u0,
        v0=v0,
        dx=dx,
        dt=dt,
        steps=steps,
        bc='dirichlet',
        source_func=source_func,
    )

    # ── Compute metrics ──────────────────────────────────────────────────
    mean_l2_error = float(np.mean(results['l2_error']))
    max_l2_error = float(np.max(results['l2_error']))
    mean_overlap = float(np.mean(results['overlap']))
    min_overlap = float(np.min(results['overlap']))

    energy_classical = results['energy_classical']
    energy_quantum = results['energy_quantum']
    energy_conservation_classical = float(np.std(energy_classical) / (np.mean(energy_classical) + 1e-30))
    energy_conservation_quantum = float(np.std(energy_quantum) / (np.mean(energy_quantum) + 1e-30))

    print("\n" + "=" * 70)
    print("VALIDATION RESULTS")
    print("=" * 70)
    print(f"L2 Error (quantum vs classical):")
    print(f"  Mean: {mean_l2_error:.6f}")
    print(f"  Max:  {max_l2_error:.6f}")
    print(f"\nQuantum Overlap |⟨ψ_q|ψ_c⟩|²:")
    print(f"  Mean: {mean_overlap:.6f}")
    print(f"  Min:  {min_overlap:.6f}")
    print(f"\nEnergy Conservation (std/mean):")
    print(f"  Classical: {energy_conservation_classical:.6e}")
    print(f"  Quantum:   {energy_conservation_quantum:.6e}")
    print("=" * 70)

    # ── Save results ─────────────────────────────────────────────────────
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = OUTPUT_DIR / timestamp
    output_dir.mkdir(parents=True, exist_ok=True)

    # JSON report
    report = {
        'timestamp': timestamp,
        'configuration': {
            'nx': nx, 'dx': dx, 'dt': dt, 'steps': steps,
            'mu_range': [mu_min, mu_max],
            'rho_range': [rho_min, rho_max],
            'source_t0': t0,
            'source_sigma': sigma_t,
        },
        'metrics': {
            'mean_l2_error': mean_l2_error,
            'max_l2_error': max_l2_error,
            'mean_overlap': mean_overlap,
            'min_overlap': min_overlap,
            'energy_conservation_classical': energy_conservation_classical,
            'energy_conservation_quantum': energy_conservation_quantum,
        },
    }

    json_path = output_dir / 'hamiltonian_validation.json'
    with open(json_path, 'w') as f:
        json.dump(report, f, indent=2)
    print(f"\nSaved JSON report: {json_path}")

    # Excel output
    excel_path = output_dir / 'hamiltonian_validation.xlsx'
    _save_excel(results, excel_path, report)
    print(f"Saved Excel: {excel_path}")

    # Plots
    FIGURE_DIR.mkdir(parents=True, exist_ok=True)
    _plot_validation(results, FIGURE_DIR)
    print(f"Saved plots to: {FIGURE_DIR}")

    print("\n✓ Hamiltonian validation complete.")


def _save_excel(results, path, report):
    """Save validation results to Excel with multiple sheets."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Sheet 1: Summary
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["Hamiltonian Validation Report"])
    ws_summary.append(["Timestamp", report['timestamp']])
    ws_summary.append([])
    ws_summary.append(["Metric", "Value"])
    for key, val in report['metrics'].items():
        ws_summary.append([key, val])

    # Sheet 2: Time series comparison
    ws_ts = wb.create_sheet("TimeSeries")
    ws_ts.append(["Time [s]", "L2 Error", "Overlap", "Energy Classical [J/m]", "Energy Quantum [J/m]"])
    for i in range(len(results['time'])):
        ws_ts.append([
            float(results['time'][i]),
            float(results['l2_error'][i]),
            float(results['overlap'][i]),
            float(results['energy_classical'][i]),
            float(results['energy_quantum'][i]),
        ])

    # Sheet 3: Configuration
    ws_config = wb.create_sheet("Configuration")
    ws_config.append(["Parameter", "Value"])
    for key, val in report['configuration'].items():
        ws_config.append([key, str(val)])

    wb.save(path)


def _plot_validation(results, figure_dir):
    """Generate validation plots."""
    time = results['time']
    l2_error = results['l2_error']
    overlap = results['overlap']
    energy_classical = results['energy_classical']
    energy_quantum = results['energy_quantum']

    # Plot 1: L2 Error
    plt.figure(figsize=(8, 5))
    plt.plot(time, l2_error, 'b-', linewidth=2)
    plt.xlabel('Time [s]', fontsize=12)
    plt.ylabel('Relative L2 Error', fontsize=12)
    plt.title('Hamiltonian Validation: ||u_quantum - u_classical|| / ||u_classical||', fontsize=13)
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(figure_dir / 'hamiltonian_validation_l2_error.png', dpi=300, bbox_inches='tight')
    plt.close()

    # Plot 2: Overlap
    plt.figure(figsize=(8, 5))
    plt.plot(time, overlap, 'g-', linewidth=2)
    plt.xlabel('Time [s]', fontsize=12)
    plt.ylabel('Quantum Overlap |⟨ψ_q|ψ_c⟩|²', fontsize=12)
    plt.title('Hamiltonian Validation: Quantum State Overlap', fontsize=13)
    plt.ylim([0, 1.05])
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(figure_dir / 'hamiltonian_validation_overlap.png', dpi=300, bbox_inches='tight')
    plt.close()

    # Plot 3: Energy comparison
    plt.figure(figsize=(8, 5))
    plt.plot(time, energy_classical, 'b-', linewidth=2, label='Classical PDE')
    plt.plot(time, energy_quantum, 'r--', linewidth=2, label='Quantum exp(-iHt)')
    plt.xlabel('Time [s]', fontsize=12)
    plt.ylabel('Total Energy [J/m]', fontsize=12)
    plt.title('Hamiltonian Validation: Energy Conservation', fontsize=13)
    plt.legend(fontsize=11)
    plt.grid(True, alpha=0.3)
    plt.tight_layout()
    plt.savefig(figure_dir / 'hamiltonian_validation_energy.png', dpi=300, bbox_inches='tight')
    plt.close()

    # Plot 4: Trajectory comparison (sample timesteps)
    traj_q = results['trajectory_quantum']
    traj_c = results['trajectory_classical']
    sample_steps = [0, len(traj_c)//4, len(traj_c)//2, 3*len(traj_c)//4, len(traj_c)-1]

    fig, axes = plt.subplots(2, 3, figsize=(15, 8))
    axes = axes.flatten()

    for idx, step in enumerate(sample_steps):
        if idx >= len(axes):
            break
        ax = axes[idx]
        u_classical = traj_c[step][1:-1]
        u_quantum = traj_q[step][1:-1]
        x_plot = np.arange(len(u_classical))

        ax.plot(x_plot, u_classical, 'b-', linewidth=2, label='Classical', marker='o')
        ax.plot(x_plot, u_quantum, 'r--', linewidth=2, label='Quantum', marker='s')
        ax.set_title(f't = {time[step]:.3f} s', fontsize=11)
        ax.set_xlabel('Grid Point', fontsize=10)
        ax.set_ylabel('Displacement u', fontsize=10)
        ax.legend(fontsize=9)
        ax.grid(True, alpha=0.3)

    # Hide unused subplot
    if len(sample_steps) < len(axes):
        axes[-1].axis('off')

    plt.suptitle('Hamiltonian Validation: Trajectory Comparison (Quantum vs Classical)', fontsize=14)
    plt.tight_layout()
    plt.savefig(figure_dir / 'hamiltonian_validation_trajectories.png', dpi=300, bbox_inches='tight')
    plt.close()


if __name__ == '__main__':
    main()
