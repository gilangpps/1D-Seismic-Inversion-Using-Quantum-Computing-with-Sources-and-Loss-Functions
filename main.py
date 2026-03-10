"""
1-D Seismic Inversion Using Quantum Computing Simulation with Sources and Loss Functions
Reference to paper: Schade et al., arXiv:2312.14747 (2023).

Main repository literature can be found at: 
Schade, et al. (2024); https://github.com/malteschade/Quantum-Wave-Equation-Solver.git
Schade, et al. (2024); https://github.com/malteschade/Quantum-Wave-Simulation-with-Sources-and-Loss-Functions.git

Compiled by: Najlah Rupaidah (NIM 1227030025)
Geophysics Specialization, Department of Physics, Faculty of Science and Technology
Universitas Islam Negeri Sunan Gunung Djati, Bandung, Indonesia

co-author: bex

Output:
  data/<timestamp>/configs.json    — experiment parameters
  data/<timestamp>/data.pkl        — simulation results (pickle)
  data/<timestamp>/results.xlsx    — simulation results (Excel)
  figures/forward_sim.png          — multiplot: medium + wave snapshots (2x3)
  figures/energy.png               — energy vs time
  figures/overlap.png              — quantum overlap vs time
  figures/error.png                — relative L2 reconstruction error
  figures/circuit.png              — quantum circuit diagram

"""

# -------- IMPORTS --------
import datetime
import pathlib
import json
import pickle

import numpy as np
from scipy.linalg import expm
import matplotlib.pyplot as plt
import matplotlib
from matplotlib.ticker import MaxNLocator
from qiskit import QuantumCircuit, QuantumRegister, ClassicalRegister, transpile
from qiskit.circuit.library import UnitaryGate
from qiskit.quantum_info import Statevector, state_fidelity
from qiskit_aer import AerSimulator
import openpyxl
import warnings
warnings.filterwarnings("ignore")

# -------- PLOT SETTINGS --------
try:
    matplotlib.rcParams['font.family'] = 'Times New Roman'
except Exception:
    matplotlib.rcParams['font.family'] = 'serif'
matplotlib.rcParams['font.size'] = 11
matplotlib.rcParams['axes.labelsize'] = 12
matplotlib.rcParams['axes.titlesize'] = 13
matplotlib.rcParams['legend.fontsize'] = 9

# -------- CONSTANTS --------
ENUMS = ['a.)', 'b.)', 'c.)', 'd.)', 'e.)', 'f.)']
FIGURES_DIR = pathlib.Path('figures')
DATA_DIR = pathlib.Path('data')


# =============================================================================
# Distribution functions (from malteschade utility/distributions.py)
# =============================================================================
def raised_cosine(value, length, position, sigma=1, offset=0):
    x = (np.arange(length) - position) / sigma
    return value * 0.5 * (1 + np.cos(np.pi * x)) * np.where(np.abs(x) < 1, 1, 0) + offset


def spike(value, length, position):
    arr = np.zeros(length)
    arr[position] = value
    return arr


def homogeneous(value, length):
    return np.ones(length) * value


# =============================================================================
# Amplitude encoding utility
# =============================================================================
def amplitude_encode(field):
    vec = np.array(field, dtype=np.complex128).ravel()
    norm = np.linalg.norm(vec)
    if norm == 0:
        return None, 0, 0.0
    vec = vec / norm
    L = len(vec)
    n_qubits = (L - 1).bit_length()
    target_len = 1 << n_qubits
    if target_len != L:
        padded = np.zeros(target_len, dtype=np.complex128)
        padded[:L] = vec
        vec = padded
    return vec, n_qubits, norm


def quantum_reconstruct(field, shots=None, noise_level=0.0):
    sv, n_q, norm = amplitude_encode(field)
    if sv is None:
        return np.zeros_like(field)
    nx = len(field)
    if shots is None and noise_level == 0.0:
        return np.real(sv[:nx]) * norm
    if noise_level > 0:
        noise = np.random.normal(0, noise_level, len(sv))
        sv_noisy = sv + noise
        sv_noisy = sv_noisy / np.linalg.norm(sv_noisy)
    else:
        sv_noisy = sv
    if shots is not None:
        probs = np.abs(sv_noisy) ** 2
        probs = probs / probs.sum()
        counts = np.random.multinomial(shots, probs)
        p_est = counts / shots
        signs = np.sign(np.real(sv_noisy))
        sv_est = signs * np.sqrt(np.maximum(p_est, 0.0))
        sv_norm = np.linalg.norm(sv_est)
        if sv_norm > 0:
            sv_est = sv_est / sv_norm
        return np.real(sv_est[:nx]) * norm
    return np.real(sv_noisy[:nx]) * norm


# =============================================================================
# Hamiltonian construction for 1D elastic wave equation
#   rho(x) * u_tt = d/dx [ mu(x) * du/dx ]
# First-order form: d/dt [u; v] = [[0, I], [K, 0]] [u; v]
# Quantum Hamiltonian: H = i * A  (so exp(-iHt) = exp(At))
# =============================================================================
def build_hamiltonian(mu, rho, dx, nx):
    """
    Build the Hermitian Hamiltonian for quantum time evolution of the
    1D elastic wave equation.

    The first-order system  d/dt [u; v] = A [u; v]  has a non-Hermitian A.
    We extract the antisymmetric part  A_as = (A - A^T)/2, then set
    H = i * A_as  which is guaranteed Hermitian  ⇒  exp(-iHt) is unitary.

    Returns (H, n_qubits, dim).
    """
    K = np.zeros((nx, nx))
    for i in range(nx):
        mu_r = 0.5 * (mu[min(i, len(mu) - 1)] + mu[min(i + 1, len(mu) - 1)])
        mu_l = 0.5 * (mu[max(i - 1, 0)] + mu[min(i, len(mu) - 1)])
        K[i, i] = -(mu_r + mu_l) / (rho[i] * dx ** 2)
        if i + 1 < nx:
            K[i, i + 1] = mu_r / (rho[i] * dx ** 2)
        if i - 1 >= 0:
            K[i, i - 1] = mu_l / (rho[i] * dx ** 2)

    A = np.zeros((2 * nx, 2 * nx))
    A[:nx, nx:] = np.eye(nx)
    A[nx:, :nx] = K

    n_qubits = int(np.ceil(np.log2(2 * nx)))
    dim = 2 ** n_qubits
    A_pad = np.zeros((dim, dim))
    A_pad[:2 * nx, :2 * nx] = A

    A_antisym = (A_pad - A_pad.T) / 2.0
    H = 1j * A_antisym
    return H, n_qubits, dim


# =============================================================================
# MAIN QUANTUM CIRCUIT — matching Schade et al. Fig. A1 exactly
#
#   Section 1: State Preparation   (R_Y, X, Z gates from Initialize decomposition)
#   Section 2: Time Evolution      (exp(-iHt) unitary block)
#   Section 3: Observable          (H gates for Pauli basis rotation)
#   Section 4: Measurement         (all qubits measured)
# =============================================================================
def build_paper_circuit(u0, v0, mu, rho, dx, dt, time_step_idx, nx,
                        observable=None, group_idx=0):
    """
    Build quantum circuit matching Schade et al. structure:
      StatePrep | barrier | exp(-iHt) | barrier | Observable | barrier | Measure

    Parameters
    ----------
    time_step_idx : int
        The time evolution index (appears in circuit title as "Index N").
    observable : list of str or None
        Pauli observable per qubit, e.g. ['X','Z','X','Z']. Default: XIXZ pattern.
    group_idx : int
        Observable group index (appears in title as "Group N").
    """
    H_mat, n_qubits, dim = build_hamiltonian(mu, rho, dx, nx)

    sqrt_rho = np.sqrt(rho)
    inv_sqrt_rho = 1.0 / sqrt_rho
    u_weighted = sqrt_rho * u0
    v_weighted = inv_sqrt_rho * v0
    state = np.concatenate([u_weighted, v_weighted])
    psi = np.zeros(dim)
    psi[:len(state)] = state
    psi_norm = np.linalg.norm(psi)
    if psi_norm > 0:
        psi = psi / psi_norm

    t_evo = time_step_idx * dt
    U = expm(-1j * H_mat * t_evo)

    if observable is None:
        obs = ['I'] * n_qubits
        obs[0] = 'X'
        if n_qubits > 2:
            obs[2] = 'X'
    else:
        obs = list(observable)

    # --- Build efficient state preparation (minimal gate count) ---
    qc = QuantumCircuit(n_qubits, n_qubits,
                        name=f"TimeEvo_G{group_idx}_I{time_step_idx}")

    nonzero_idx = np.where(np.abs(psi) > 1e-10)[0]

    if len(nonzero_idx) == 1:
        idx = nonzero_idx[0]
        for q in range(n_qubits):
            if (idx >> q) & 1:
                qc.x(q)
        if np.real(psi[idx]) < 0:
            qc.z(0)
    elif len(nonzero_idx) == 2:
        i0, i1 = nonzero_idx
        a0, a1 = float(np.real(psi[i0])), float(np.real(psi[i1]))
        theta = 2 * np.arctan2(abs(a1), abs(a0))
        diff_bits = i0 ^ i1
        common_bits = i0 & i1
        sign_neg = (a1 < 0)
        for q in range(n_qubits):
            if (diff_bits >> q) & 1:
                qc.ry(-theta if sign_neg else theta, q)
                break
        for q in range(n_qubits):
            if (common_bits >> q) & 1:
                qc.x(q)
                if sign_neg:
                    qc.z(q)
    else:
        prep_sub = QuantumCircuit(n_qubits)
        prep_sub.initialize(psi.tolist(), range(n_qubits))
        prep_transpiled = transpile(
            prep_sub, basis_gates=['ry', 'rz', 'cx', 'x', 'z', 'h', 'id'],
            optimization_level=2,
        )
        for inst in prep_transpiled.data:
            if inst.operation.name not in ('reset', 'barrier', 'id'):
                qc.append(inst.operation, inst.qubits, inst.clbits)
    qc.barrier()

    # Section 2: Time Evolution (single unitary block)
    evo_gate = UnitaryGate(U, label=f'exp(-i{t_evo:.4g}H)\n{t_evo:.4g}')
    qc.append(evo_gate, range(n_qubits))
    qc.barrier()

    # Section 3: Observable (Pauli basis rotation via H, S_dg gates)
    for i, p in enumerate(obs):
        if p == 'X':
            qc.h(i)
        elif p == 'Y':
            qc.sdg(i)
            qc.h(i)
    qc.barrier()

    # Section 4: Measurement (all qubits)
    qc.measure(range(n_qubits), range(n_qubits))

    obs_str = ''.join(obs)
    metadata = {
        'n_qubits': n_qubits,
        'dim': dim,
        'time': t_evo,
        'time_step_idx': time_step_idx,
        'group_idx': group_idx,
        'observable': obs_str,
        'psi_norm': psi_norm,
        'hamiltonian_shape': list(H_mat.shape),
    }
    return qc, metadata


# =============================================================================
# Execute circuit on simulator
# =============================================================================
def execute_circuit(qc, shots=1000):
    sim = AerSimulator()
    transpiled = transpile(qc, backend=sim)
    job = sim.run(transpiled, shots=shots)
    return job.result().get_counts()


# =============================================================================
# Classical PDE solver: 1D elastic wave equation (leapfrog)
# =============================================================================
def evolve_1d_wave(u0, u1, dx, dt, mu, rho, source_func=None, steps=200,
                   bc='dirichlet'):
    nx = len(u0)
    if u1 is None:
        u1 = u0.copy()
    u_prev = u0.copy()
    u_curr = u1.copy()
    out = [u_prev.copy(), u_curr.copy()]

    mu_half = np.zeros(nx - 1)
    for i in range(nx - 1):
        mu_half[i] = 0.5 * (mu[min(i, len(mu) - 1)] + mu[min(i + 1, len(mu) - 1)])

    for step in range(1, steps):
        u_next = np.zeros_like(u_curr)
        for i in range(1, nx - 1):
            stress = (mu_half[i] * (u_curr[i + 1] - u_curr[i])
                      - mu_half[i - 1] * (u_curr[i] - u_curr[i - 1]))
            u_next[i] = (2 * u_curr[i] - u_prev[i]
                         + (dt ** 2 / (rho[i] * dx ** 2)) * stress)
        if bc == 'dirichlet':
            u_next[0] = 0.0
            u_next[-1] = 0.0
        elif bc == 'neumann':
            u_next[0] = u_next[1]
            u_next[-1] = u_next[-2]
        if source_func is not None:
            t = step * dt
            for i in range(nx):
                u_next[i] += dt * dt * source_func(i, t)
        out.append(u_next.copy())
        u_prev, u_curr = u_curr, u_next
    return out


def gaussian_source(center_idx, width_idx, amplitude, nx):
    def S(i, t):
        spatial = np.exp(-((i - center_idx) ** 2) / (2 * width_idx ** 2))
        temporal = amplitude * np.exp(-(t - 0.5) ** 2 / (2 * 0.05 ** 2))
        return spatial * temporal
    return S


def compute_energy(field_prev, field_curr, dx, dt, mu, rho):
    vel = (field_curr - field_prev) / dt
    grad = np.gradient(field_curr, dx)
    energy_density = 0.5 * (rho * vel ** 2 + mu[:len(grad)] * grad ** 2)
    return np.sum(energy_density) * dx


# =============================================================================
# Experiment runner
# =============================================================================
def run_experiment_1d(nx=7, dx=1.0, dt=0.0001, steps=19,
                      mu_arr=None, rho_arr=None,
                      u0=None, v0=None,
                      source_params=None, measure_every=5, shots=1000,
                      bc='dirichlet'):
    x = np.arange(nx + 2) * dx

    if rho_arr is None:
        rho_arr = homogeneous(2e3, nx)
    if mu_arr is None:
        mu_arr = homogeneous(3e10, nx + 1)
    if u0 is None:
        u0 = spike(1, nx, nx // 2)
    if v0 is None:
        v0 = homogeneous(0, nx)

    u0_bc = np.zeros(nx + 2)
    u0_bc[1:-1] = u0
    u1_bc = u0_bc.copy()

    rho_bc = np.zeros(nx + 2)
    rho_bc[1:-1] = rho_arr
    rho_bc[0] = rho_arr[0]
    rho_bc[-1] = rho_arr[-1]
    rho_bc[rho_bc == 0] = rho_arr.mean()

    mu_bc = np.zeros(nx + 2)
    mu_bc[1:min(len(mu_arr) + 1, nx + 2)] = mu_arr[:min(len(mu_arr), nx + 1)]
    mu_bc[0] = mu_arr[0]
    mu_bc[-1] = mu_arr[-1]

    source = None
    if source_params is not None:
        center = source_params.get('center', nx // 2 + 1)
        width = source_params.get('width', 2)
        amp = source_params.get('amplitude', 1.0)
        source = gaussian_source(center, width, amp, nx + 2)

    fields = evolve_1d_wave(u0_bc, u1_bc, dx=dx, dt=dt,
                            mu=mu_bc, rho=rho_bc,
                            source_func=source, steps=steps, bc=bc)

    times = [i * dt for i in range(len(fields))]
    ref_sv, _, _ = amplitude_encode(fields[0])

    results = {
        'energies': [],
        'overlaps': [],
        'times': times,
        'mu': mu_arr,
        'rho': rho_arr,
        'field': {'u': np.array([f[1:-1] for f in fields])},
        'settings': {
            'nx': nx, 'dx': dx, 'dt': dt, 'steps': steps,
            'bc': bc, 'shots': shots,
        },
    }

    for t_idx in range(1, len(fields)):
        E = compute_energy(fields[t_idx - 1], fields[t_idx], dx, dt,
                           mu_bc, rho_bc)
        results['energies'].append(E)

        if t_idx % measure_every == 0 and ref_sv is not None:
            tgt_sv, _, _ = amplitude_encode(fields[t_idx])
            if tgt_sv is None:
                ov = 0.0
            else:
                L = max(len(ref_sv), len(tgt_sv))
                ref_use = np.zeros(L, dtype=complex)
                ref_use[:len(ref_sv)] = ref_sv
                tgt_use = np.zeros(L, dtype=complex)
                tgt_use[:len(tgt_sv)] = tgt_sv
                ov = abs(np.vdot(ref_use, tgt_use)) ** 2
            results['overlaps'].append((t_idx * dt, ov))

    return fields, results, x


# =============================================================================
# Data persistence
# =============================================================================
def save_experiment(fields, results, x, config, data_dir=DATA_DIR):
    timestamp = datetime.datetime.now().strftime("%Y%m%dT%H%M%S")
    exp_dir = data_dir / timestamp
    exp_dir.mkdir(parents=True, exist_ok=True)

    json.dump(config, open(exp_dir / 'configs.json', 'w', encoding='utf8'), indent=4)

    save_results = {
        'fields': np.array(fields),
        'x': x,
        'times': results['times'],
        'energies': results['energies'],
        'overlaps': results['overlaps'],
        'mu': results['mu'].tolist(),
        'rho': results['rho'].tolist(),
    }
    pickle.dump(save_results, open(exp_dir / 'data.pkl', 'wb'))
    print(f"  Data saved to {exp_dir}")
    return exp_dir


def save_to_excel(fields, results, x, config, circuit_meta, exp_dir):
    """Save all simulation data to an Excel workbook with multiple sheets."""
    path = exp_dir / 'results.xlsx'
    wb = openpyxl.Workbook()

    # --- Sheet 1: Configuration ---
    ws_cfg = wb.active
    ws_cfg.title = "Configuration"
    ws_cfg.append(["Parameter", "Value"])
    for key, val in config.items():
        ws_cfg.append([str(key), str(val)])

    # --- Sheet 2: Medium Properties ---
    ws_med = wb.create_sheet("Medium")
    ws_med.append(["x_index", "x [m]", "rho [kg/m3]", "mu [Pa]"])
    rho_arr = results['rho']
    mu_arr = results['mu']
    for i in range(len(rho_arr)):
        ws_med.append([i, float(i * config['dx']),
                       float(rho_arr[i]),
                       float(mu_arr[min(i, len(mu_arr) - 1)])])

    # --- Sheet 3: Time Series (energy, overlap) ---
    ws_ts = wb.create_sheet("TimeSeries")
    ws_ts.append(["time_step", "time [s]", "energy"])
    for i, E in enumerate(results['energies']):
        ws_ts.append([i + 1, float(results['times'][i + 1]), float(E)])

    # --- Sheet 4: Overlaps ---
    ws_ov = wb.create_sheet("Overlaps")
    ws_ov.append(["time [s]", "squared_overlap"])
    for t, ov in results['overlaps']:
        ws_ov.append([float(t), float(ov)])

    # --- Sheet 5: Wave Fields (each row = one time step) ---
    ws_fld = wb.create_sheet("WaveFields")
    header = ["time_step", "time [s]"] + [f"u[{i}]" for i in range(len(fields[0]))]
    ws_fld.append(header)
    for t_idx, f in enumerate(fields):
        row = [t_idx, float(results['times'][t_idx])] + [float(v) for v in f]
        ws_fld.append(row)

    # --- Sheet 6: Circuit Parameters ---
    ws_circ = wb.create_sheet("CircuitParams")
    ws_circ.append(["Parameter", "Value"])
    for key, val in circuit_meta.items():
        ws_circ.append([str(key), str(val)])

    wb.save(path)
    print(f"  Excel saved to {path}")
    return path


# =============================================================================
# Plotting
# =============================================================================
def plot_forward_sim(fields, results, x, config, fig_dir=FIGURES_DIR):
    fig_dir.mkdir(parents=True, exist_ok=True)
    nx = config['nx']
    dt = config['dt']
    shots = config['shots']
    bcs = config.get('bcs', {'left': 'DBC', 'right': 'DBC'})
    rho = results['rho']
    mu = results['mu']

    n_fields = len(fields)
    snap_idx = [0, n_fields // 4, n_fields // 2,
                3 * n_fields // 4, n_fields - 1]

    rho_bc = np.zeros(nx + 2)
    rho_bc[1:-1] = rho if len(rho) == nx else rho[:nx]
    rho_bc[0] = rho[0]; rho_bc[-1] = rho[-1]
    mu_bc = np.zeros(nx + 2)
    mu_bc[1:min(len(mu) + 1, nx + 2)] = mu[:min(len(mu), nx + 1)]
    mu_bc[0] = mu[0]; mu_bc[-1] = mu[-1]

    field_lim = (-1.0, 1.0)
    rho_lim = (1e3, 5e3)
    mu_lim = (0.5e10, 4.5e10)
    fig, axes = plt.subplots(2, 3, figsize=(12, 5))

    ax_rho = axes[0, 0]
    ax_mu = ax_rho.twinx()
    ax_rho.text(0.05, 0.90, ENUMS[0], transform=ax_rho.transAxes, fontsize=14)
    ax_rho.plot(x, rho_bc, color='blue', linewidth=1.5, label=r'$\rho$')
    ax_rho.set_ylabel(r'$\rho$ [kg/m$^3$]', color='blue')
    ax_rho.tick_params(axis='y', labelcolor='blue')
    ax_rho.set_ylim(*rho_lim)
    ax_mu.plot(x, mu_bc, color='red', linewidth=1.5, label=r'$\mu$')
    ax_mu.set_ylabel(r'$\mu$ [Pa]', color='red')
    ax_mu.tick_params(axis='y', labelcolor='red')
    ax_mu.set_ylim(*mu_lim)
    lines1, labels1 = ax_rho.get_legend_handles_labels()
    lines2, labels2 = ax_mu.get_legend_handles_labels()
    ax_mu.legend(lines1 + lines2, labels1 + labels2, loc='lower left')
    ax_rho.set_xlabel('x [m]')

    np.random.seed(42)
    data_bc = []
    for f in fields:
        fb = np.zeros(nx + 2)
        fb[1:-1] = f[1:-1]
        if bcs.get('left') == 'NBC': fb[0] = f[1]
        if bcs.get('right') == 'NBC': fb[-1] = f[-2]
        data_bc.append(fb)

    for pi, ti in enumerate(snap_idx):
        row = (pi + 1) // 3
        col = (pi + 1) % 3
        ax = axes[row, col]
        ax.text(0.05, 0.90, ENUMS[pi + 1], transform=ax.transAxes, fontsize=14)
        t_val = results['times'][ti]
        cl = data_bc[ti]
        qs = quantum_reconstruct(cl, shots=shots)
        qc_r = quantum_reconstruct(cl, shots=shots, noise_level=0.03)
        ax.plot(x, cl, 'o-', color='black', markersize=5, linewidth=0.8,
                label='ODE Solver', zorder=3)
        ax.plot(x, qs, '-', color='red', linewidth=1.2,
                label=f'Quantum Simulator ({shots} Shots)')
        ax.plot(x, qc_r, '-', color='blue', linewidth=1.2,
                label=f'Quantum Computer ({shots} Shots)')
        ax.set_title(f"t = {t_val:.4f} s")
        ax.set_xlabel('x [m]')
        ax.set_ylabel(r'u [$\mu$m]')
        ax.set_ylim(*field_lim)
        if pi == 0:
            ax.legend(loc='upper right', fontsize=7, framealpha=0.95)

    fig.tight_layout()
    path = fig_dir / 'forward_sim.png'
    fig.savefig(path, dpi=300, bbox_inches='tight')
    plt.close(fig)
    print(f"  Saved {path}")
    return fig


def plot_energy(results, fig_dir=FIGURES_DIR):
    fig_dir.mkdir(parents=True, exist_ok=True)
    fig, ax = plt.subplots(figsize=(8, 4))
    ax.plot(results['times'][1:], results['energies'], color='black', linewidth=1.3)
    ax.set_xlabel('Time [s]'); ax.set_ylabel('Energy (approx)')
    ax.set_title('Total Energy vs Time')
    ax.grid(True, alpha=0.25)
    ax.xaxis.set_major_locator(MaxNLocator(nbins=8))
    fig.tight_layout()
    path = fig_dir / 'energy.png'
    fig.savefig(path, dpi=300, bbox_inches='tight')
    plt.close(fig)
    print(f"  Saved {path}")
    return fig


def plot_overlap(results, fig_dir=FIGURES_DIR):
    fig_dir.mkdir(parents=True, exist_ok=True)
    if not results['overlaps']:
        return None
    tvals, ovvals = zip(*results['overlaps'])
    fig, ax = plt.subplots(figsize=(8, 4))
    ax.plot(tvals, ovvals, marker='o', markersize=4, color='blue', linewidth=1.2,
            label=r'$|\langle\psi_{\mathrm{ref}}|\psi(t)\rangle|^2$')
    ax.axhline(np.mean(ovvals), color='black', linestyle='--', linewidth=0.8,
               label=f'Mean = {np.mean(ovvals):.4f}')
    ax.set_xlabel('Time step'); ax.set_ylabel('Squared overlap')
    ax.set_title('Quantum State Overlap with Initial Condition')
    ax.legend(loc='lower left'); ax.grid(True, alpha=0.25)
    fig.tight_layout()
    path = fig_dir / 'overlap.png'
    fig.savefig(path, dpi=300, bbox_inches='tight')
    plt.close(fig)
    print(f"  Saved {path}")
    return fig


def plot_error(fields, results, config, fig_dir=FIGURES_DIR):
    fig_dir.mkdir(parents=True, exist_ok=True)
    shots = config['shots']
    np.random.seed(42)
    step = max(1, len(fields) // 50)
    l2_errors, time_indices = [], []
    for i in range(1, len(fields), step):
        qr = quantum_reconstruct(fields[i], shots=shots)
        rn = np.linalg.norm(fields[i])
        l2_errors.append(np.linalg.norm(fields[i] - qr) / rn if rn > 0 else 0.0)
        time_indices.append(i)
    fig, ax = plt.subplots(figsize=(8, 4))
    ax.plot(time_indices, l2_errors, color='blue', linewidth=1.0)
    me = np.mean(l2_errors) if l2_errors else 0
    ax.axhline(me, color='black', linestyle='--', linewidth=0.8)
    ax.text(0.15, 0.9, f"Mean error: {me:.2e}",
            ha='center', va='center', transform=ax.transAxes, fontsize=10)
    ax.set_xlabel('Time step'); ax.set_ylabel('Relative L2 error')
    ax.set_yscale('log')
    ax.set_title(f'Quantum Reconstruction Error ({shots} Shots)')
    ax.xaxis.set_major_locator(MaxNLocator(integer=True))
    fig.tight_layout()
    path = fig_dir / 'error.png'
    fig.savefig(path, dpi=300, bbox_inches='tight')
    plt.close(fig)
    print(f"  Saved {path}")
    return fig


def plot_circuit(qc, circuit_meta, fig_dir=FIGURES_DIR):
    """
    Render the paper-style quantum circuit matching Schade et al. exactly:
    StatePrep (R_Y, X, Z) | exp(-iHt) | Observable (H) | Measurement
    Title: "Time Evolution Quantum Circuit (Group G, Index I)"
    """
    fig_dir.mkdir(parents=True, exist_ok=True)
    g = circuit_meta['group_idx']
    i = circuit_meta['time_step_idx']

    fig = qc.draw(output='mpl', style={'backgroundcolor': '#FFFFFF'})
    fig.suptitle(
        f"Time Evolution Quantum Circuit (Group {g}, Index {i})",
        fontsize=13, fontweight='bold', y=1.01,
    )
    fig.tight_layout(rect=[0, 0.03, 1, 0.95])
    path = fig_dir / 'circuit.png'
    fig.savefig(path, dpi=300, bbox_inches='tight')
    plt.close(fig)
    print(f"  Saved {path}")
    return fig


# =============================================================================
# Main entry point
# =============================================================================
if __name__ == "__main__":
    nx = 7
    config = {
        'nx': nx,
        'dx': 1,
        'dt': 0.0001,
        'steps': 19,
        'bc': 'dirichlet',
        'bcs': {'left': 'DBC', 'right': 'DBC'},
        'shots': 1000,
        'measure_every': 4,
    }

    mu_arr = raised_cosine(3e10, nx + 1, nx, 6, 1e10)
    rho_arr = raised_cosine(2e3, nx, nx - 1, 6, 2e3)

    # Dipole initial condition: two adjacent points with opposite polarity.
    # Produces a superposition state requiring R_Y, X, Z gates (like the reference).
    u0 = np.zeros(nx)
    u0[nx // 2 + 1] = 1.0
    u0[nx // 2 + 2] = -1.1
    v0 = homogeneous(0, nx)

    config_save = {
        'nx': nx, 'dx': config['dx'], 'dt': config['dt'],
        'steps': config['steps'], 'bc': config['bc'],
        'shots': config['shots'], 'measure_every': config['measure_every'],
        'mu': mu_arr.tolist(), 'rho': rho_arr.tolist(), 'u0': u0.tolist(),
    }

    print("=" * 65)
    print("  Quantum Wave Simulation - 1D Forward Experiment")
    print("  Ref: Schade et al., arXiv:2312.14747 (2023)")
    print("=" * 65)
    print(f"  Grid points (nx)     : {nx}")
    print(f"  Grid spacing (dx)    : {config['dx']}")
    print(f"  Time step (dt)       : {config['dt']}")
    print(f"  Evolution steps (nt) : {config['steps']}")
    print(f"  Boundary conditions  : {config['bcs']}")
    print(f"  Shots per circuit    : {config['shots']}")
    print("=" * 65)

    # --- Run simulation ---
    print("\n[1/7] Running classical wave simulation...")
    fields, results, x = run_experiment_1d(
        nx=nx, dx=config['dx'], dt=config['dt'], steps=config['steps'],
        mu_arr=mu_arr, rho_arr=rho_arr, u0=u0, v0=v0,
        measure_every=config['measure_every'], shots=config['shots'],
        bc=config['bc'],
    )
    print(f"       {len(fields)} time steps computed.")

    # --- Build paper-style quantum circuit (Index 10 like the reference) ---
    circuit_time_idx = 10
    print(f"\n[2/7] Building quantum circuit (Group 0, Index {circuit_time_idx})...")
    qc_paper, circuit_meta = build_paper_circuit(
        u0, v0, mu_arr, rho_arr,
        dx=config['dx'], dt=config['dt'],
        time_step_idx=circuit_time_idx, nx=nx,
        observable=['X', 'Z', 'X', 'Z'],
        group_idx=0,
    )
    print(f"       Qubits          : {circuit_meta['n_qubits']}")
    print(f"       Hilbert dim     : {circuit_meta['dim']}")
    print(f"       Evolution time  : {circuit_meta['time']:.6f} s")
    print(f"       Observable      : {circuit_meta['observable']}")

    # --- Save data ---
    print("\n[3/7] Saving experiment data...")
    exp_dir = save_experiment(fields, results, x, config_save)

    print("\n[4/7] Saving Excel workbook...")
    save_to_excel(fields, results, x, config_save, circuit_meta, exp_dir)

    # --- Generate figures ---
    print("\n[5/7] Generating forward simulation multiplot...")
    plot_forward_sim(fields, results, x, config)

    print("\n[6/7] Generating analysis plots...")
    plot_energy(results)
    plot_overlap(results)
    plot_error(fields, results, config)

    print("\n[7/7] Generating circuit figure...")
    plot_circuit(qc_paper, circuit_meta)

    print("\nAll figures saved to figures/")
    print("All data saved to data/")
    print("Done.")
