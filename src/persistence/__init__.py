"""
src/persistence/__init__.py
Excel workbook and file persistence for 1-D Seismic Inversion.

Sheets (Requirement J):
  1. Configuration      — experiment parameters
  2. Medium             — x, rho, mu
  3. TimeSeries         — time, classical PDE energy
  4. Overlaps           — time, squared quantum overlap |<ψ_rec|ψ_ref>|²
  5. WaveFields         — full displacement field per time step
  6. CircuitParams      — quantum circuit metadata
  7. Source             — source amplitude vs position
  8. Loss               — per-timestep INVERSION misfit J(t; μ_rec)
                          (NOT quantum reconstruction MSE)
  9. ModelUpdate        — mu_initial vs mu_updated vs mu_true
  10. OptimizationHistory — iteration, J(μ), mean overlap, μ per iteration

Additional sheet (if quantum_recon_loss present):
  11. QuantumReconLoss  — per-timestep quantum reconstruction MSE (encoding diagnostic)

Change log vs prior version
-----------------------------
* Sheet 8 'Loss' now stores the per-timestep INVERSION misfit
  results['loss'] = J(t; μ_rec) = (1/N)‖u_fwd(t;μ_rec)−u_ref(t)‖²
  Previously it stored quantum reconstruction MSE, which is NOT the
  scientific objective of the inversion.
* Column header changed from 'reconstruction_MSE' to 'inversion_misfit_MSE'.
* Added optional Sheet 11 'QuantumReconLoss' for the reconstruction diagnostic.
"""

import datetime
import json
import pickle

import numpy as np
import openpyxl

from src.constants import DATA_DIR


def save_experiment(fields, results, x, config, data_dir=DATA_DIR):
    """Save experiment fields and metadata to JSON + pickle."""
    timestamp = datetime.datetime.now().strftime("%Y%m%dT%H%M%S")
    exp_dir   = data_dir / timestamp
    exp_dir.mkdir(parents=True, exist_ok=True)

    # JSON config
    json.dump(
        config,
        open(exp_dir / 'configs.json', 'w', encoding='utf8'),
        indent=4, default=str,
    )

    # Pickle results
    save_results = {
        'fields':   [f.tolist() for f in fields],
        'x':        x.tolist() if hasattr(x, 'tolist') else list(x),
        'times':    results['times'],
        'energies': results['energies'],
        'overlaps': results['overlaps'],
        'mu':       (results['mu'].tolist()
                     if hasattr(results['mu'], 'tolist')
                     else list(results['mu'])),
        'rho':      (results['rho'].tolist()
                     if hasattr(results['rho'], 'tolist')
                     else list(results['rho'])),
    }
    pickle.dump(save_results, open(exp_dir / 'data.pkl', 'wb'))
    print(f"  Data saved to {exp_dir}")
    return exp_dir


def save_to_excel(fields, results, x, config, circuit_meta, exp_dir):
    """
    Save all simulation data to Excel with required sheets.

    Parameters
    ----------
    fields : list of np.ndarray
        Wavefield snapshots from the RECOVERED-model forward simulation.
    results : dict
        Must contain:
            'times', 'energies', 'overlaps', 'mu', 'rho',
            'loss'              — per-timestep inversion misfit J(t;μ_rec)
            'source_amplitude'  — spatial source profile
            'mu_initial', 'mu_updated', 'mu_true' (optional)
            'optimization_history' — opt_results dict from SeismicOptimizer
        May optionally contain:
            'quantum_recon_loss'     — recon MSE on reference fields
            'quantum_recon_loss_rec' — recon MSE on recovered-model fields
    """
    path = exp_dir / 'results.xlsx'
    wb   = openpyxl.Workbook()

    # ── Sheet 1: Configuration ────────────────────────────────────────────
    ws = wb.active
    ws.title = "Configuration"
    ws.append(["Parameter", "Value"])
    for k, v in config.items():
        ws.append([str(k), str(v)])

    # ── Sheet 2: Medium ───────────────────────────────────────────────────
    ws_med = wb.create_sheet("Medium")
    ws_med.append(["x_index", "x [m]", "rho [kg/m3]", "mu [Pa]"])
    rho_arr = results['rho']
    mu_arr  = results['mu']
    for i in range(len(rho_arr)):
        ws_med.append([
            i,
            float(i * config.get('dx', 1.0)),
            float(rho_arr[i]),
            float(mu_arr[min(i, len(mu_arr) - 1)]),
        ])

    # ── Sheet 3: TimeSeries ───────────────────────────────────────────────
    ws_ts = wb.create_sheet("TimeSeries")
    ws_ts.append(["time_step", "time [s]", "classical_PDE_energy [J/m]"])
    for i, E in enumerate(results['energies']):
        ws_ts.append([i + 1, float(results['times'][i + 1]), float(E)])

    # ── Sheet 4: Overlaps ─────────────────────────────────────────────────
    # Overlap = |<ψ_fwd(t; μ_rec) | ψ_ref(t; μ_true)>|²
    # Inversion quality metric: 0 for a wrong model, 1 for a perfect model.
    ws_ov = wb.create_sheet("Overlaps")
    ws_ov.append(["time [s]", "squared_overlap |<ψ_rec|ψ_ref>|^2"])
    for t, ov in results.get('overlaps', []):
        ws_ov.append([float(t), float(ov)])

    # ── Sheet 5: WaveFields ───────────────────────────────────────────────
    ws_fld = wb.create_sheet("WaveFields")
    n_pts  = len(fields[0]) if fields else 0
    header = ["time_step", "time [s]"] + [f"u[{i}]" for i in range(n_pts)]
    ws_fld.append(header)
    for t_idx, f in enumerate(fields):
        row = [t_idx, float(results['times'][t_idx])] + [float(v) for v in f]
        ws_fld.append(row)

    # ── Sheet 6: CircuitParams ────────────────────────────────────────────
    ws_circ = wb.create_sheet("CircuitParams")
    ws_circ.append(["Parameter", "Value"])
    for k, v in circuit_meta.items():
        ws_circ.append([str(k), str(v)])

    # ── Sheet 7: Source ───────────────────────────────────────────────────
    if 'source_amplitude' in results:
        ws_src = wb.create_sheet("Source")
        ws_src.append(["x_index", "x [m]", "source_amplitude"])
        src = results['source_amplitude']
        for i in range(len(src)):
            xi = (float(x[i])
                  if i < len(x)
                  else float(i * config.get('dx', 1.0)))
            ws_src.append([i, xi, float(src[i])])

    # ── Sheet 8: Loss — per-timestep INVERSION misfit ────────────────────
    # J(t; μ_rec) = (1/N) ‖u_fwd(t;μ_rec) − u_ref(t;μ_true)‖²
    # This is the correct scientific loss from the inversion, NOT the
    # quantum reconstruction MSE.
    if 'loss' in results:
        ws_loss = wb.create_sheet("Loss")
        ws_loss.append(["time_step", "time [s]", "inversion_misfit_MSE"])
        times    = results['times']
        loss_arr = results['loss']
        for i, lv in enumerate(loss_arr):
            t_i = (float(times[i + 1])
                   if i + 1 < len(times)
                   else float(i * config.get('dt', 1.0)))
            ws_loss.append([i + 1, t_i, float(lv)])

    # ── Sheet 9: ModelUpdate ──────────────────────────────────────────────
    if 'mu_initial' in results and 'mu_updated' in results:
        ws_mu   = wb.create_sheet("ModelUpdate")
        mu_init = results['mu_initial']
        mu_upd  = results['mu_updated']
        mu_true = results.get('mu_true', [None] * len(mu_init))
        ws_mu.append(["index", "mu_initial [Pa]", "mu_updated [Pa]", "mu_true [Pa]"])
        for i in range(len(mu_init)):
            mt = (float(mu_true[i])
                  if i < len(mu_true) and mu_true[i] is not None
                  else '')
            ws_mu.append([i, float(mu_init[i]), float(mu_upd[i]), mt])

    # ── Sheet 10: OptimizationHistory ────────────────────────────────────
    # Stores J(μ) at every Adam iteration together with mean overlap and μ.
    # This is the authoritative inversion history from SeismicOptimizer.
    opt = results.get('optimization_history', {})
    if opt:
        ws_opt    = wb.create_sheet("OptimizationHistory")
        loss_hist = opt.get('loss_history', [])
        iter_hist = opt.get('iteration_history', [])
        mu_hist   = opt.get('mu_history', [])
        ov_hist   = opt.get('overlaps_history', [])

        n_mu = len(mu_hist[0]) if mu_hist else 0
        ws_opt.append(
            ["iteration", "inversion_loss_J(mu)", "mean_overlap_|<psi_rec|psi_ref>|^2"]
            + [f"mu[{i}] [Pa]" for i in range(n_mu)]
        )

        for idx in range(len(iter_hist)):
            it   = iter_hist[idx] if idx < len(iter_hist) else idx
            loss = float(loss_hist[idx]) if idx < len(loss_hist) else ''
            ovs  = ov_hist[idx] if idx < len(ov_hist) else []
            mean_ov = (float(np.mean([o for _, o in ovs]))
                       if ovs else '')
            mu_row  = ([float(v) for v in mu_hist[idx]]
                       if idx < len(mu_hist) else [])
            ws_opt.append([it, loss, mean_ov] + mu_row)

    # ── Sheet 11: QuantumReconLoss (optional diagnostic) ─────────────────
    # Per-timestep quantum reconstruction MSE on reference fields.
    # Encoding quality metric: how well amplitude encoding + shot measurement
    # can reconstruct the classical wavefield.
    # This is DISTINCT from the inversion misfit in Sheet 8.
    qrl = results.get('quantum_recon_loss', [])
    if qrl:
        ws_qrl = wb.create_sheet("QuantumReconLoss")
        ws_qrl.append([
            "time_step", "time [s]",
            "recon_MSE_reference",
            "recon_MSE_recovered",
        ])
        times     = results['times']
        qrl_rec   = results.get('quantum_recon_loss_rec', [None] * len(qrl))
        for i, mse_ref in enumerate(qrl):
            t_i = (float(times[i + 1])
                   if i + 1 < len(times)
                   else float(i * config.get('dt', 1.0)))
            mse_rec = float(qrl_rec[i]) if i < len(qrl_rec) and qrl_rec[i] is not None else ''
            ws_qrl.append([i + 1, t_i, float(mse_ref), mse_rec])

    wb.save(path)
   # ── Sheet 12: PerformanceMetrics ────────────────────────────────
    perf = results.get("performance_metrics", {})

    if perf:
        ws_perf = wb.create_sheet("PerformanceMetrics")

        ws_perf.append(["Metric", "Value"])

        for key, value in perf.items():
            ws_perf.append([key, value])

    # ── Sheet 13: HamiltonianValidation (optional) ───────────────────────
    # Independent quantum exp(-iHt) vs classical leapfrog comparison
    # Validates that Hamiltonian H = i(A − Aᵀ)/2 captures wave physics
    # This is DISTINCT from:
    #   - Sheet 4 (Overlaps): inversion quality |⟨ψ_rec|ψ_ref⟩|²
    #   - Sheet 11 (QuantumReconLoss): encoding quality diagnostic
    h_val = results.get('hamiltonian_validation', None)
    if h_val is not None:
        ws_hval = wb.create_sheet("HamiltonianValidation")
        ws_hval.append([
            "time [s]",
            "L2_error (quantum vs classical)",
            "state_overlap |⟨ψ_q|ψ_c⟩|²",
            "classical_energy [J/m]",
            "quantum_state_norm",
        ])
        time_arr = h_val['time']
        l2_err = h_val['l2_error']
        overlap = h_val['overlap']
        energy = h_val['energy_classical']
        norm_q = h_val['norm_quantum']
        for i in range(len(time_arr)):
            ws_hval.append([
                float(time_arr[i]),
                float(l2_err[i]),
                float(overlap[i]),
                float(energy[i]),
                float(norm_q[i]),
            ])

    wb.save(path)
    print(f"  Excel saved to {path}")
    return path
